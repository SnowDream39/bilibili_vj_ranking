import os
from openpyxl import load_workbook
from openpyxl.cell import MergedCell

old_values = {"奶昔": "奶昔PurinN","Raindonut_0511":"Raindo0511", "でるた(Δ)":"でるた"}  # 需要替换的键值对

target_column = 'author'  # 例如 'synthesizer' 代表只在该列查找，None 表示全表
partial_replace = False  # True: 进行部分匹配替换, False: 只进行精确匹配

def process_excel_file(file_path: str):
    try:
        wb = load_workbook(file_path)
        modified = False  

        for ws in wb.worksheets:
            col_idx = None
            if target_column:
                for col in ws.iter_cols(min_row=1, max_row=1):
                    if col[0].value == target_column:
                        col_idx = col[0].column
                        break

            for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row,
                                    min_col=ws.min_column, max_col=ws.max_column):
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    if col_idx and cell.column != col_idx:
                        continue  

                    for old_value, new_value in old_values.items():
                        if partial_replace and isinstance(cell.value, str) and old_value in cell.value:
                            print(f"文件 {file_path} 工作表 {ws.title} 单元格 {cell.coordinate}：'{cell.value}' -> '{cell.value.replace(old_value, new_value)}'")
                            cell.value = cell.value.replace(old_value, new_value)
                            modified = True
                        elif cell.value == old_value:
                            print(f"文件 {file_path} 工作表 {ws.title} 单元格 {cell.coordinate}：'{old_value}' -> '{new_value}'")
                            cell.value = new_value
                            modified = True

        if modified:
            wb.save(file_path)
        else:
            print(f"文件 {file_path} 中没有匹配的内容")
    except Exception as e:
        print(f"处理文件 {file_path} 时出错：{e}")

def traverse_folder(root_folder: str):
    for dirpath, dirnames, filenames in os.walk(root_folder):
        for filename in filenames:
            if filename.lower().endswith(".xlsx"):
                file_path = os.path.join(dirpath, filename)
                process_excel_file(file_path)

if __name__ == "__main__":
    root_folder = os.path.abspath(os.path.dirname(__file__))
    traverse_folder(root_folder)
