# utils/io_utils.py
# IO工具模块，提供文件保存和数据格式化等通用功能。
import pandas as pd
from pathlib import Path
from typing import Optional, List, Union
from openpyxl.utils import get_column_letter
from utils.logger import logger

def save_to_excel(df: pd.DataFrame, filename: Union[str, Path], usecols: Optional[List[str]] = None):
    """
    保存DataFrame到Excel文件
       
    Args:
        df (pd.DataFrame): 要保存的DataFrame。
        filename (Union[str, Path]): 保存路径（字符串或Path对象）。
        usecols (Optional[List[str]], optional): 指定要保存的列名列表。
    """
    if usecols:
        # 仅保留在DataFrame中存在的指定列
        cols_to_save = [col for col in usecols if col in df.columns]
        df = df[cols_to_save].copy() 
    try:
        # 将'aid'列转换为正整数的字符串格式，以防科学计数法
        if 'aid' in df.columns:
            df['aid'] = df['aid'].apply(lambda x: "{:.0f}".format(float(x)) if pd.notna(x) and str(x).strip() != '' else '')

        # 使用ExcelWriter写入数据并设置格式
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            
            # 设置'pubdate'列为文本格式并左对齐
            if 'pubdate' in df.columns:
                col_idx = df.columns.get_loc('pubdate')
                if isinstance(col_idx, int):
                    pubdate_col_letter = get_column_letter(col_idx + 1)
                    for cell in worksheet[pubdate_col_letter]:
                        cell.number_format = '@'
                        cell.alignment = cell.alignment.copy(horizontal='left')
            
            # 设置'aid'列为文本格式并左对齐
            if 'aid' in df.columns:
                col_idx = df.columns.get_loc('aid')
                if isinstance(col_idx, int):
                    aid_col_letter = get_column_letter(col_idx + 1)
                    for cell in worksheet[aid_col_letter]:
                        cell.number_format = '@'
                        cell.alignment = cell.alignment.copy(horizontal='left')
                    
        logger.info(f"{filename} 保存完成")
    except Exception as e:
        logger.warning(f"Excel 保存失败：{e}")

        # 如果Excel保存失败，则备份为CSV文件
        backup_csv = Path(filename).with_suffix('.csv')
        df.to_csv(backup_csv, index=False, encoding='utf-8-sig')
        logger.info(f"数据已备份至 {backup_csv}")

def format_columns(df):
    """将DataFrame中指定的数值列格式化为保留两位小数的字符串。

    Args:
        df (pd.DataFrame): 要处理的DataFrame。
        
    Returns:
        pd.DataFrame: 处理后的DataFrame。
    """
    columns = ['viewR', 'favoriteR', 'coinR', 'likeR', 'fixA', 'fixB', 'fixC']
    for col in columns:
        if col in df.columns:
            # 将列转换为数值类型，无法转换的值设为NaN
            df[col] = pd.to_numeric(df[col], errors='coerce')
            # 格式化数值:保留2位小数，NaN值转换为空字符串
            df[col] = df[col].apply(lambda x: f'{x:.2f}' if pd.notnull(x) else '')
    return df